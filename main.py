import openpyxl

# telecharge le fichier et le lit a l'aide de openpy.load_workbook("nom_fichier_a_lire")
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# le dictionnaire exercice 1 des produits par fournisseur
product_per_supplier = {}
# dictionnaire exercice 2 valeur total par fournisseur
total_value_per_supplier = {}
# dictionnaire exercice 3 valeurs des produits < 10
products_under_10_inv = {}

# affiche le nombre maximal des rangees de la liste de produit dans le fichier iventory.xlsx
# print(product_list.max_row)

for product_row in range(2, product_list.max_row + 1):
    # liste de produits dans chaque cellule de la rangee(colonne 1) et la colonne 4
    supplier_name = product_list.cell(product_row, 4).value
    # dans inventory.xlsx on prend inventory qui est a la colonne 2
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value

    # ajout de la colonne 5 dans inventory et calcul
    inventory_price = product_list.cell(product_row,5)

    # Exercie 1: calcul du nombre de produits par fournisseur
    # si le nom du fournisseur(supplier_name) existe dans produit par fournisseur(product_per_supplier)
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier[supplier_name]
        # current_num_products +1
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    # Exercice 2: calcul de la valeur totale de chaque fournisseur dans le fichier inventory
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Exercice 3: produits dans l'inventory < 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Exercice 4: ajout de la valeur et calcul du prix total dans inventory
    inventory_price.value = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
print(inventory_price)

# sauvegarde des ajouts dans un nouveau fichier
inv_file.save("inventory with total value.xlsx")
